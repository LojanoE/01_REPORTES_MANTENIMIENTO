"""
Migracion de datos Excel -> Supabase para modulo de Bodegas (v3.0)
Lee el archivo Excel y carga inventario + transacciones en Supabase.

Uso:
    python migrate_bodegas.py              # Migracion normal (upsert)
    python migrate_bodegas.py --clean      # Limpia datos existentes antes de cargar
    python migrate_bodegas.py --dry-run    # Solo valida sin insertar
"""

import openpyxl
from supabase import create_client, Client
from datetime import datetime, time as dt_time
import sys
import logging
from dataclasses import dataclass, field

# ===========================================================================
# CONFIGURACION
# ===========================================================================

SUPABASE_URL = "https://dzmhhlsttqygjvfabdxx.supabase.co"
SUPABASE_KEY = "sb_publishable_gqH1ZQ9bl--GBVfmcI4Q-A__UvBKemK"
EXCEL_PATH = r"BODEGAS\1. Control_Bodegas_V0.xlsm"
BATCH_SIZE = 100

# ===========================================================================
# LOGGING
# ===========================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("bodegas_migration")


# ===========================================================================
# STATS TRACKER
# ===========================================================================

@dataclass
class MigrationStats:
    inventory_read: int = 0
    inventory_inserted: int = 0
    inventory_skipped: int = 0
    ingresos_read: int = 0
    ingresos_valid: int = 0
    ingresos_skipped: int = 0
    egresos_read: int = 0
    egresos_valid: int = 0
    egresos_skipped: int = 0
    transactions_inserted: int = 0
    transactions_failed: int = 0
    errors: list = field(default_factory=list)

    def summary(self) -> str:
        lines = [
            "=" * 60,
            "RESUMEN DE MIGRACION",
            "=" * 60,
            f"Inventario leido:      {self.inventory_read}",
            f"Inventario insertado:  {self.inventory_inserted}",
            f"Inventario omitido:    {self.inventory_skipped}",
            f"Ingresos leidos:       {self.ingresos_read}",
            f"Ingresos validos:      {self.ingresos_valid}",
            f"Ingresos omitidos:     {self.ingresos_skipped}",
            f"Egresos leidos:        {self.egresos_read}",
            f"Egresos validos:       {self.egresos_valid}",
            f"Egresos omitidos:      {self.egresos_skipped}",
            f"Transacciones OK:      {self.transactions_inserted}",
            f"Transacciones fallidas:{self.transactions_failed}",
        ]
        if self.errors:
            lines.append(f"\nERRORES ({len(self.errors)}):")
            for e in self.errors[:20]:
                lines.append(f"  - {e}")
            if len(self.errors) > 20:
                lines.append(f"  ... y {len(self.errors) - 20} errores mas")
        lines.append("=" * 60)
        return "\n".join(lines)


# ===========================================================================
# HELPERS DE LIMPIEZA
# ===========================================================================

def clean(val: any) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if s in ("-", "nan", "None", "#VALUE!", "#REF!", ""):
        return ""
    return s


def clean_num(val: any) -> float:
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def parse_time(val: any) -> str | None:
    if val is None:
        return None
    if isinstance(val, dt_time):
        return val.isoformat()
    s = str(val).strip()
    if ":" in s:
        parts = s.split(":")
        try:
            return f"{int(parts[0]):02d}:{int(parts[1]):02d}:00"
        except (ValueError, IndexError):
            return None
    return None


def parse_date(val: any) -> str:
    if val is None:
        return datetime.now().isoformat()[:10]
    if isinstance(val, datetime):
        return val.isoformat()[:10]
    s = str(val).strip()
    if not s or s in ("-", "nan", "None", "#VALUE!", "#REF!"):
        return datetime.now().isoformat()[:10]
    for fmt in ("%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.isoformat()[:10]
        except ValueError:
            continue
    try:
        parts = s.replace("/", "-").split("-")
        if len(parts) == 3:
            if len(parts[2]) == 2:
                parts[2] = "20" + parts[2]
            day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
            dt = datetime(year, month, day)
            return dt.isoformat()[:10]
    except (ValueError, IndexError):
        pass
    if len(s) >= 10:
        return s[:10]
    return datetime.now().isoformat()[:10]


# ===========================================================================
# SUPABASE CLIENT
# ===========================================================================

def get_supabase() -> Client:
    return create_client(SUPABASE_URL, SUPABASE_KEY)


# ===========================================================================
# CLEAN RELOAD
# ===========================================================================

def clean_reload(supabase: Client):
    logger.info("Limpiando datos existentes (TRUNCATE)...")
    try:
        supabase.rpc("exec_sql", {"sql": "TRUNCATE bodegas_transactions RESTART IDENTITY CASCADE"}).execute()
        logger.info("  bodegas_transactions: limpiada")
    except Exception:
        pass

    try:
        supabase.table("bodegas_transactions").delete().neq("id", 0).execute()
        logger.info("  bodegas_transactions: limpiada (delete)")
    except Exception as e:
        logger.warning(f"  Error limpiando transactions: {e}")

    try:
        supabase.table("bodegas_inventory").delete().neq("id", 0).execute()
        logger.info("  bodegas_inventory: limpiada")
    except Exception as e:
        logger.warning(f"  Error limpiando inventory: {e}")

    logger.info("Limpieza completada")


# ===========================================================================
# MIGRACION
# ===========================================================================

def migrate(do_clean: bool = False, dry_run: bool = False):
    stats = MigrationStats()
    supabase = get_supabase()

    logger.info(f"Cargando Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    if do_clean and not dry_run:
        clean_reload(supabase)

    # =======================================================================
    # 1. INVENTARIO
    # =======================================================================
    logger.info("\n[1/3] Procesando INVENTARIO...")
    ws_inv = wb["Inventario"]

    items = []
    seen_ids = set()

    for row_idx in range(7, ws_inv.max_row + 1):
        cod_unico = clean(ws_inv.cell(row_idx, 3).value)
        if not cod_unico or cod_unico == "COD_UNICO":
            continue

        if cod_unico in seen_ids:
            stats.inventory_skipped += 1
            continue
        seen_ids.add(cod_unico)

        stats.inventory_read += 1

        code = clean(ws_inv.cell(row_idx, 2).value)
        if code:
            code = code.split(".")[0]

        description = clean(ws_inv.cell(row_idx, 4).value)
        if not description:
            description = cod_unico

        item = {
            "excel_unique_id": cod_unico,
            "code": code or cod_unico,
            "description": description,
            "specifications": clean(ws_inv.cell(row_idx, 5).value),
            "unit": clean(ws_inv.cell(row_idx, 7).value) or "UND",
            "initial_stock": clean_num(ws_inv.cell(row_idx, 8).value),
            "current_stock": clean_num(ws_inv.cell(row_idx, 11).value),
            "min_stock": 0,
            "bodega": clean(ws_inv.cell(row_idx, 14).value),
            "location": clean(ws_inv.cell(row_idx, 15).value),
        }
        items.append(item)

    if items and not dry_run:
        logger.info(f"  Subiendo {len(items)} items en lotes de {BATCH_SIZE}...")
        for i in range(0, len(items), BATCH_SIZE):
            batch = items[i:i + BATCH_SIZE]
            try:
                supabase.table("bodegas_inventory").upsert(
                    batch, on_conflict="excel_unique_id"
                ).execute()
                stats.inventory_inserted += len(batch)
                logger.info(f"  Lote {i // BATCH_SIZE + 1}: {len(batch)} items")
            except Exception as e:
                msg = f"Error lote inventario {i}: {e}"
                logger.error(f"  {msg}")
                stats.errors.append(msg)

    elif dry_run:
        logger.info(f"  [DRY RUN] {len(items)} items listos para subir")

    stats.inventory_inserted = len(items) if dry_run else stats.inventory_inserted

    # Mapeo excel_unique_id -> db_id
    logger.info("  Construyendo mapeo de IDs...")
    res = supabase.table("bodegas_inventory").select("id, excel_unique_id").execute()
    uid_map = {item["excel_unique_id"]: item["id"] for item in res.data}
    logger.info(f"  Mapeo: {len(uid_map)} items")

    # =======================================================================
    # 2. INGRESOS
    # =======================================================================
    logger.info("\n[2/3] Procesando INGRESOS...")
    ws_in = wb["Ingresos"]

    transactions_in = []
    for row_idx in range(7, ws_in.max_row + 1):
        stats.ingresos_read += 1
        cod_unico = clean(ws_in.cell(row_idx, 7).value)
        if not cod_unico or cod_unico not in uid_map:
            stats.ingresos_skipped += 1
            continue

        qty = clean_num(ws_in.cell(row_idx, 11).value)
        if qty <= 0:
            stats.ingresos_skipped += 1
            continue

        stats.ingresos_valid += 1
        transactions_in.append({
            "item_id": uid_map[cod_unico],
            "type": "IN",
            "quantity": qty,
            "date": parse_date(ws_in.cell(row_idx, 3).value),
            "time": parse_time(ws_in.cell(row_idx, 4).value),
            "bodega": clean(ws_in.cell(row_idx, 6).value),
            "specifications": clean(ws_in.cell(row_idx, 9).value),
            "received_by": clean(ws_in.cell(row_idx, 21).value),
            "dispatched_by": clean(ws_in.cell(row_idx, 22).value),
            "voucher_code": clean(ws_in.cell(row_idx, 18).value),
            "location": clean(ws_in.cell(row_idx, 23).value),
            "notes": clean(ws_in.cell(row_idx, 24).value),
            "created_by_name": clean(ws_in.cell(row_idx, 22).value) or clean(ws_in.cell(row_idx, 21).value) or "Migracion",
        })

    logger.info(f"  {stats.ingresos_valid} validos, {stats.ingresos_skipped} omitidos")

    # =======================================================================
    # 3. EGRESOS
    # =======================================================================
    logger.info("\n[3/3] Procesando EGRESOS...")
    ws_out = wb["Egresos"]

    transactions_out = []
    for row_idx in range(6, ws_out.max_row + 1):
        stats.egresos_read += 1
        cod_unico = clean(ws_out.cell(row_idx, 6).value)
        if not cod_unico or cod_unico not in uid_map:
            stats.egresos_skipped += 1
            continue

        qty = clean_num(ws_out.cell(row_idx, 10).value)
        if qty <= 0:
            stats.egresos_skipped += 1
            continue

        stats.egresos_valid += 1
        transactions_out.append({
            "item_id": uid_map[cod_unico],
            "type": "OUT",
            "quantity": qty,
            "date": parse_date(ws_out.cell(row_idx, 2).value),
            "time": parse_time(ws_out.cell(row_idx, 3).value),
            "bodega": clean(ws_out.cell(row_idx, 5).value),
            "specifications": clean(ws_out.cell(row_idx, 8).value),
            "received_by": "",
            "dispatched_by": clean(ws_out.cell(row_idx, 12).value),
            "voucher_code": "",
            "location": clean(ws_out.cell(row_idx, 13).value),
            "notes": clean(ws_out.cell(row_idx, 14).value),
            "created_by_name": clean(ws_out.cell(row_idx, 11).value),
        })

    logger.info(f"  {stats.egresos_valid} validos, {stats.egresos_skipped} omitidos")

    # =======================================================================
    # 4. SUBIR TRANSACCIONES
    # =======================================================================
    all_trans = transactions_in + transactions_out
    logger.info(f"\nTotal transacciones a subir: {len(all_trans)}")

    if all_trans and not dry_run:
        for i in range(0, len(all_trans), BATCH_SIZE):
            batch = all_trans[i:i + BATCH_SIZE]
            try:
                supabase.table("bodegas_transactions").insert(batch).execute()
                stats.transactions_inserted += len(batch)
                logger.info(f"  Lote {i // BATCH_SIZE + 1}: {len(batch)} transacciones")
            except Exception as e:
                logger.warning(f"  Lote {i // BATCH_SIZE + 1} fallo, insertando uno a uno...")
                for t in batch:
                    try:
                        supabase.table("bodegas_transactions").insert([t]).execute()
                        stats.transactions_inserted += 1
                    except Exception as e2:
                        stats.transactions_failed += 1
                        if stats.transactions_failed <= 10:
                            err = f"Fila {t.get('type')} {t.get('date')} item_id={t.get('item_id')}: {e2}"
                            logger.error(f"    {err}")
                            stats.errors.append(err)

    elif dry_run:
        logger.info(f"  [DRY RUN] {len(all_trans)} transacciones listas")
        stats.transactions_inserted = len(all_trans)

    # =======================================================================
    # 5. RESUMEN
    # =======================================================================
    print(f"\n{stats.summary()}")

    if stats.transactions_failed > 0:
        logger.warning(f"Migracion completada con {stats.transactions_failed} errores")
        sys.exit(1)
    else:
        logger.info("Migracion completada exitosamente")


# ===========================================================================
# MAIN
# ===========================================================================

if __name__ == "__main__":
    clean_flag = "--clean" in sys.argv
    dry_run_flag = "--dry-run" in sys.argv

    if clean_flag:
        logger.info("Modo CLEAN: se limpiaran datos existentes antes de cargar")
    if dry_run_flag:
        logger.info("Modo DRY RUN: no se insertaran datos")

    migrate(do_clean=clean_flag, dry_run=dry_run_flag)
